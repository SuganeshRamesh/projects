import sys, re
import hashlib
import hmac
import openpyxl
from openpyxl import Workbook
import logging
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QComboBox, QPushButton, QTableView, QHeaderView, QLabel,
                             QLineEdit, QSplitter, QFormLayout, QMessageBox, QCompleter,
                             QDialog, QDialogButtonBox, QDoubleSpinBox, QStyleFactory,QWizard,QWizardPage,QGridLayout)
from PyQt6.QtCore import QSortFilterProxyModel, Qt, QAbstractTableModel, pyqtSignal
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QPalette, QColor
from PyQt6.QtCore import QAbstractListModel, Qt
from PyQt6.QtCore import QItemSelectionModel
from PyQt6.QtWidgets import QInputDialog

class PandasModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data
        self._sort_column = None
        self._sort_order = Qt.SortOrder.AscendingOrder

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            return str(self._data.iloc[index.row(), index.column()])
        return None

    def headerData(self, section, orientation, role):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return str(self._data.columns[section])
            if orientation == Qt.Orientation.Vertical:
                return str(self._data.index[section])
        return None

    def sort(self, column, order):
        self.layoutAboutToBeChanged.emit()
        self._sort_column = self._data.columns[column]
        self._sort_order = order

        try:
            self._data = self._data.sort_values(
                by=self._sort_column,
                ascending=(order == Qt.SortOrder.AscendingOrder),
                na_position='last'
            )
        except TypeError:
            self._data = self._data.astype(str).sort_values(
                by=self._sort_column,
                ascending=(order == Qt.SortOrder.AscendingOrder),
                na_position='last'
            )

        self.layoutChanged.emit()

class StringListModel(QAbstractListModel):
    def __init__(self, strings, parent=None):
        super().__init__(parent)
        self.strings = strings

    def data(self, index, role):
        if role == Qt.ItemDataRole.DisplayRole:
            return self.strings[index.row()]
        return None

    def rowCount(self, parent=None):
        return len(self.strings)

class FilterWidget(QWidget):
    filtersApplied = pyqtSignal(dict)
    reset_all_filters = pyqtSignal()
    resetFilters = pyqtSignal()
    def __init__(self, parent=None, columns=None, data=None):
        super().__init__(parent)
        self.setMinimumWidth(300)
        layout = QFormLayout(self)
        self.filters = {}
        self.data = data
        self.columns = columns
        self.labels = {}  # Store references to labels
        self.highlighted_label = None

        for col in columns:
            if pd.api.types.is_numeric_dtype(data[col]):
                value_combo = QComboBox(self)
                value_combo.setEditable(True)
                value_combo.setPlaceholderText("Value ")
                value_combo.addItems([str(value) for value in data[col].unique()])

                if col in ['drill dia', 'corner radius','Point Dia','d1', 'd2', 'd3','D1','D2','Cutting Dia','cutter dia']:
                    range_combo = QComboBox(self)
                    range_combo.setEditable(True)
                    range_combo.setPlaceholderText("Tolerance")
                    range_combo.addItems([f"±{i/100:.2f}" for i in range(5, 105, 5)])
                elif col in ['flute length', 'Flute Length', 'flute Length', 'L', 'L1', 'L2', 'L3']:
                    range_combo = QComboBox(self)
                    range_combo.setEditable(True)
                    range_combo.setPlaceholderText("Tolerance")
                    range_combo.addItems([f"±{i}" for i in range(1, 21)])
                else:
                    range_combo = None

                if range_combo is not None:
                    combo_layout = QHBoxLayout()
                    combo_layout.addWidget(value_combo)
                    tolerance_label = QLabel("Tolerance:")
                    combo_layout.addWidget(tolerance_label)
                    combo_layout.addWidget(range_combo)

                    label = QLabel(f"{col}:")
                    layout.addRow(label, combo_layout)
                    self.filters[col] = (value_combo, range_combo)
                    self.labels[col] = label
                    value_combo.currentTextChanged.connect(lambda text, c=col: self.update_filters(c))
                    range_combo.currentTextChanged.connect(lambda text, c=col: self.update_filters(c))
                    value_combo.lineEdit().textChanged.connect(lambda text, c=col: self.highlight_label(c, True))
                    range_combo.lineEdit().textChanged.connect(lambda text, c=col: self.highlight_label(c, True))
                    value_combo.lineEdit().editingFinished.connect(lambda c=col: self.highlight_label(c, False))
                    range_combo.lineEdit().editingFinished.connect(lambda c=col: self.highlight_label(c, False))
                    value_combo.currentIndexChanged.connect(lambda index, c=col: self.highlight_label(c, True))
                    range_combo.currentIndexChanged.connect(lambda index, c=col: self.highlight_label(c, True))
                    value_combo.activated.connect(lambda checked, c=col: self.highlight_label(c, True))
                    range_combo.activated.connect(lambda checked, c=col: self.highlight_label(c, True))
                    value_combo.setStyleSheet(self.get_combobox_style())
                    range_combo.setStyleSheet(self.get_combobox_style())
                else:
                    label = QLabel(f"{col}:")
                    layout.addRow(label, value_combo)
                    self.filters[col] = value_combo
                    self.labels[col] = label
                    value_combo.currentTextChanged.connect(lambda text, c=col: self.update_filters(c))
                    value_combo.lineEdit().textChanged.connect(lambda text, c=col: self.highlight_label(c, True))
                    value_combo.lineEdit().editingFinished.connect(lambda c=col: self.highlight_label(c, False))
                    value_combo.setStyleSheet(self.get_combobox_style())

            else:
                combo_box = QComboBox(self)
                combo_box.setEditable(True) 
                combo_box.setPlaceholderText("Select Value")
                unique_values = self.get_sorted_unique_values(data[col])
                unique_values = [str(value) for value in unique_values]  # Convert all values to strings
                combo_box.addItems([str(value) for value in unique_values])

                # **Autocomplete Setup**
                completer = QCompleter(unique_values, combo_box)  # Use unique_values directly
                completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                completer.setFilterMode(Qt.MatchFlag.MatchContains)  # Match anywhere in the string
                combo_box.setCompleter(completer)

                label = QLabel(col)
                layout.addRow(label, combo_box)
                self.filters[col] = combo_box
                self.labels[col] = label
                combo_box.currentTextChanged.connect(lambda text, c=col: self.update_filters(c))
                combo_box.lineEdit().textChanged.connect(lambda text, c=col: self.highlight_label(c, True))
                combo_box.lineEdit().editingFinished.connect(lambda c=col: self.highlight_label(c, False))

                combo_box.setStyleSheet(self.get_combobox_style())

        self.reset_button = QPushButton("Reset Filters", self)
        self.reset_button.clicked.connect(self.reset_filters)
        layout.addRow(self.reset_button)

    def get_combobox_style(self):
        return """
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 15px;
                border-left-width: 1px;
                border-left-color: #3a3a3a;
                border-left-style: solid;
                image: url('arrow_down.png');  # Replace with your arrow image path
            }
        """
    def highlight_label(self, col, highlight):
        if self.highlighted_label:
            self.highlighted_label.setStyleSheet("")
        label = self.labels.get(col)
        if label:
            if highlight:
                label.setStyleSheet("background-color: #00FFF9; color: black;")  # Highlight color
                self.highlighted_label = label
            else:
                label.setStyleSheet("")
                self.highlighted_label = None

    def get_sorted_unique_values(self, series):
        unique_values = series.dropna().unique().tolist()
        return sorted(unique_values, key=lambda x: str(x).lower())

    def update_filters(self, changed_col):
        current_criteria = self.get_filter_criteria()
        filtered_data = self.apply_criteria_to_data(current_criteria)

        for col in self.columns:
            widget = self.filters[col]
            if isinstance(widget, tuple):
                value_combo, range_combo = widget
                unique_values = self.get_sorted_unique_values(filtered_data[col])
                current_value = value_combo.currentText()
                current_range = range_combo.currentText()

                value_combo.blockSignals(True)
                value_combo.clear()
                value_combo.addItems([str(value) for value in unique_values])
                if current_value in [str(value_combo.itemText(i)) for i in range(value_combo.count())]:
                    value_combo.setCurrentText(current_value)
                elif current_value:
                    value_combo.addItem(current_value)
                    value_combo.setCurrentText(current_value)
                else:
                    value_combo.setCurrentIndex(-1)
                value_combo.blockSignals(False)

                range_combo.blockSignals(True)
                range_combo.setCurrentText(current_range)
                range_combo.blockSignals(False)

                # Highlight label if either combo box has a value
                self.highlight_label(col, bool(current_value) or bool(current_range))
            else:
                unique_values = self.get_sorted_unique_values(filtered_data[col])
                current_value = widget.currentText()

                widget.blockSignals(True)
                widget.clear()
                widget.addItems([str(value) for value in unique_values])
                if current_value in [str(widget.itemText(i)) for i in range(widget.count())]:
                    widget.setCurrentText(current_value)
                elif current_value:
                    widget.addItem(current_value)
                    widget.setCurrentText(current_value)
                else:
                    widget.setCurrentIndex(-1)
                widget.blockSignals(False)

                # Highlight label if the combo box has a value
                self.highlight_label(col, bool(current_value))

        self.apply_filters()

    def apply_filters(self):
        criteria = self.get_filter_criteria()
        self.filtersApplied.emit(criteria)

    def get_filter_criteria(self):
        criteria = {}
        for col, widget in self.filters.items():
            if isinstance(widget, tuple):  # For drill dia, flute length, and corner radius
                value_combo, range_combo = widget
                value = value_combo.currentText()
                range_value = range_combo.currentText()
                if value or range_value:
                    criteria[col] = (value, range_value)
            else:
                value = widget.currentText()
                if value:
                    criteria[col] = value
        return criteria

    def apply_criteria_to_data(self, criteria):
        filtered_data = self.data.copy()
        for col, value in criteria.items():
            if isinstance(value, tuple):  # For drill dia, flute length, and corner radius
                selected_value, selected_range = value
                if selected_value:
                    try:
                        selected_value = float(selected_value)
                        if selected_range:
                            range_value = float(selected_range.replace('±', ''))
                            filtered_data = filtered_data[(filtered_data[col] >= selected_value - range_value) &
                                                          (filtered_data[col] <= selected_value + range_value)]
                        else:
                            filtered_data = filtered_data[filtered_data[col] == selected_value]
                    except ValueError:
                        continue
            elif value:
                filtered_data = filtered_data[filtered_data[col].astype(str).str.lower() == str(value).lower()]
        return filtered_data

    def reset_filters(self):
        for col, widget in self.filters.items():
            if isinstance(widget, tuple):
                value_combo, range_combo = widget
                value_combo.setCurrentIndex(-1)
                range_combo.setCurrentIndex(-1)
            else:
                widget.setCurrentIndex(-1)
            self.highlight_label(col, False)
        self.update_filters(None)
        self.resetFilters.emit()  # Emit the signal
        self.reset_all_filters.emit()  # Emit the signal to reset all filters



class AddToolDialog(QDialog):
    def __init__(self, dataframes, parent=None, tool_data=None):
        super().__init__(parent)
        self.dataframes = dataframes
        self.tool_data = tool_data
        self.setWindowTitle("Add Tool")
        self.setGeometry(100, 100, 1200, 800)
        self.setWindowFlags(
            self.windowFlags() |  # Keep existing flags 
            Qt.WindowType.WindowMinimizeButtonHint | 
            Qt.WindowType.WindowMaximizeButtonHint
        )
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Tool selection
        tool_layout = QHBoxLayout()
        self.tool_combo = QComboBox()
        self.tool_combo.addItems(["Select Tool"] + list(self.dataframes.keys()))
        self.tool_combo.currentTextChanged.connect(self.update_fields)
        tool_layout.addWidget(QLabel("Tool:"))
        tool_layout.addWidget(self.tool_combo)
        layout.addLayout(tool_layout)

        # Split view
        split_layout = QHBoxLayout()
        
        # Left side - input fields
        self.input_layout = QVBoxLayout()
        self.inputs = {}
        self.input_widget = QWidget()
        self.input_widget.setLayout(self.input_layout)
        self.input_widget.setFixedWidth(300)  # Increased width for better visibility
        split_layout.addWidget(self.input_widget)

        self.table_view = QTableView()
        self.table_view.setSortingEnabled(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_view.setHorizontalScrollMode(QTableView.ScrollMode.ScrollPerPixel)
        self.table_view.setSortingEnabled(True)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)  # Select entire rows
        self.table_view.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self.table_view.setEditTriggers(QTableView.EditTrigger.NoEditTriggers)  # Prevent editing of values
        self.table_view.setStyleSheet("background-color: white; color: black;")
        self.table_view.horizontalHeader().setStyleSheet("background-color: #2a2a2a; color: white;")
        self.table_view.verticalHeader().setStyleSheet("background-color: #2a2a2a; color: white;")

        # Do not set the model here
        self.model = QStandardItemModel()
        self.proxy_model = QSortFilterProxyModel()
        self.proxy_model.setSourceModel(self.model)
        self.table_view.setModel(self.proxy_model)
        # self.table_view.setModel(self.model)
        split_layout.addWidget(self.table_view)
        
        layout.addLayout(split_layout)

        # Add button
        self.add_button = QPushButton("Add Tool")
        self.add_button.clicked.connect(self.add_tool)
        layout.addWidget(self.add_button)
        
        if self.tool_data is not None:
            self.tool_combo.setCurrentText(list(self.dataframes.keys())[0])
            self.update_fields()
            for col, value in self.tool_data.items():
                if col in self.inputs:
                    self.inputs[col].setCurrentText(str(value))


    def highlight_row(self):
        selection = self.table_view.selectionModel().selectedRows()
        if selection:
            row = selection[0].row()
            self.table_view.scrollTo(self.model.index(row, 0))  # Scroll to the selected row
            self.table_view.selectionModel().select(self.model.index(row, 0), QItemSelectionModel.SelectionFlag.Select)

    def update_fields(self):
            tool = self.tool_combo.currentText()
            if tool != "Select Tool":
                df = self.dataframes[tool]
                
                # Clear existing inputs
                for i in reversed(range(self.input_layout.count())): 
                    self.input_layout.itemAt(i).widget().setParent(None)
                self.inputs.clear()

                # Add new inputs
                for col in df.columns:
                    combo = QComboBox()
                    combo.setEditable(True)
                    combo.setFixedHeight(23)  # Increased height for better visibility
                    combo.setFixedWidth(275)  # Increased width for better visibility
                    combo.addItems([str(val) for val in df[col].unique()])
                    combo.setCurrentText("")
                    combo.currentTextChanged.connect(self.filter_table)
                    self.input_layout.addWidget(QLabel(col))
                    self.input_layout.addWidget(combo)
                    self.inputs[col] = combo

                # Set values for input fields if tool_data is provided
                if self.tool_data:
                    for col, value in self.tool_data.items():
                        if col in self.inputs:
                            self.inputs[col].setCurrentText(str(value))

                # Set up table view
                self.model = QStandardItemModel()  # Set the model here
                self.model.setHorizontalHeaderLabels(df.columns)
                for _, row in df.iterrows():
                    items = [QStandardItem(str(cell)) for cell in row]
                    self.model.appendRow(items)

                self.proxy_model = QSortFilterProxyModel()
                self.proxy_model.setSourceModel(self.model)
                self.table_view.setModel(self.proxy_model)

                # Connect to the selectionChanged signal after setting the model
                self.table_view.selectionModel().selectionChanged.connect(self.highlight_row)  # Highlight selected row
                # Set last stock number
                last_stock_number = df['Stock Number'].iloc[-1]
                self.inputs['Stock Number'].setCurrentText(self .increment_stock_number(last_stock_number))


    def filter_table(self):
        tool = self.tool_combo.currentText()
        df = self.dataframes.get(tool, None)

        if df is not None:
            filtered_df = df.copy()

            for column, combo in self.inputs.items():
                if column == "Stock Number":
                    continue

                text = combo.currentText()
                if text:
                    if pd.api.types.is_numeric_dtype(df[column]):
                        try:
                            selected_value = float(text)
                            filtered_df = filtered_df[filtered_df[column] == selected_value]
                        except ValueError:
                            continue 
                    else:
                        # Only apply .str.lower() if the column has string data
                        filtered_df = filtered_df[filtered_df[column].astype(str).str.lower() == text.lower()]
 

            # Update the table view
            self.model = QStandardItemModel()
            self.model.setHorizontalHeaderLabels(filtered_df.columns)
            for _, row in filtered_df.iterrows():
                items = [QStandardItem(str(cell)) for cell in row]
                self.model.appendRow(items)

            self.proxy_model.setSourceModel(self.model)
            self.table_view.setModel(self.proxy_model)

        else:
            QMessageBox.warning(self, "Warning", f"No data available for {tool}.") 
    
    def increment_stock_number(self, stock_number):
        prefix = ''.join(filter(str.isalpha, stock_number))
        number_part = ''.join(filter(str.isdigit, stock_number))
        
        # Handle cases where there's no numeric part
        if not number_part:
            return None

        # Convert the numeric part to an integer and increment
        number = int(number_part)
        return f"{prefix}{number + 1}"  # Adjust padding if needed

    def add_tool(self):
        tool = self.tool_combo.currentText()
        if tool == "Select Tool":
            QMessageBox.warning(self, "Warning", "Please select a tool.")
            return

        # Get the new data entered by the user
        new_data = {col: combo.currentText() for col, combo in self.inputs.items() if combo.currentText()}

        # Validate and convert to float where necessary
        try:
            for col in ['drill dia', 'cutter dia', 'shank dia', 'flute length', 'corner radius', 'Tip angle', 'Shank Dia', 'Corner Radius']:
                if col in new_data and new_data[col] != '':  # Ensure the value is not an empty string
                    new_data[col] = float(new_data[col])  # Convert to float if necessary
        except ValueError:
            QMessageBox.warning(self, "Warning", "Invalid data type for one or more fields.")
            return

        # Prompt the user to select the material group
        matl_group, ok = QInputDialog.getItem(self, "Select Material Group", "Material Group:", ["ELECTRODE", "TOOL(PUR)", "TOOLS SPL", "TOOLFIX"], 0, False)
        if not ok:
            return
        
        # Add the material group to the new_data dictionary
        #new_data['Matl Group'] = matl_group

        # Define a dictionary to store key columns for each tool (sheet)
        # Dictionary with key columns for each tool
        tool_key_columns = {
            "Drills carbide": ['Point Dia', 'Shank Dia', 'Length'],
            "Drills HSS": ['Point Dia', 'Shank Dia'],
            "Special Centre Drills": ['D', 'da', 'Angle'],
            "Insert type Drills": ['Dia'],
            "Drills": ['drill dia', 'shank dia', 'flute length', 'Tip angle'],
            "FL Drill": ['Flute length', 'drill dia', 'Shank dia', 'Tip angle'],
            "Countersink Drills": ['d1', 'd2', 'd3', 'Angle'],
            "GUN Drill": ['D1', 'D2', 'L1', 'L2', 'L3'],
            "Special Drill": ['Drill Dia','Shank Dia'],
            "SPAD Drill": ['drill dia', 'Shank dia', 'L'],
            "Ball End Mills": ['cutter dia', 'Shank Dia', 'flute length', 'Helix angle'],
            "Flat End Mill": ['cutter dia', 'flute length', 'shank dia', 'Corner Radius'],
            "Insert Type Cutter": ['ORDERING CODE'],
            "Toric End Mill": ['cutter dia', 'flute length', 'corner radius', 'Shank Dia'],
            "Taper End Mill": ['cutter dia', 'Flute length', 'Shank Dia'],
            "Slitting Wheel(Carbide)": ['Bore Dia', 'Cutting Dia'],
            "Chamfer Cutter(Carbide)": ['TOOL DIA', 'INCLUDE ANGLE'],
            "Slitting Wheel(HSS)": ['Bore Dia', 'Cutting Dia'],
            "Special End Mill": ['cutter dia', 'flute length', 'shank dia'],
            "Batch End Mill": ['cutter dia', 'flute length', 'tip angle'],
            "T Slot Cutters": ['cutter dia', 'flute length', 'Shank Dia'],
            "Thread Form Cutters": ['pitch'],
            "Thread Mill Relife Cutters": ['cutter dia', 'tip angle'],
            "HandTaps-MetricCoarse&Fineserie": ['pitch'],
            "HandTaps UNC Series": ['TPI'],
            "Hand Taps UNF Series": ['TPI'],
            "Hand Taps UNJF": ['TPI'],
            "Hand Taps Metric Fine Series": ['pitch'],
            "Machine Taps Metric J Series": ['pitch'],
            "Machine Taps UNJC,UNJF,UNJEF": ['TPI'],
            "Machine Taps UNF & NF Series": ['pitch'],
            "Machine Taps UNEF Series": ['pitch'],
            "Machine Taps UNC Series": ['TPI'],
            "Machine Taps UNS Type": ['pitch'],
            "Helicoil Taps Metric&UN Series": ['pitch'],
            "Helicoil Inserts Metric & UN": ['pitch'],
            "Spiral Lock Taps": ['pitch'],
            "HeliCoil Machine Taps": ['pitch'],
            "Reamers": ['Cutter dia', 'Shank Dia']
        }

        # Replace or integrate this dictionary in your `add_tool` function as needed


        # Get the key columns for the current tool, if they exist
        key_columns = tool_key_columns.get(tool, [])

        df = self.dataframes[tool]

        def match_column(df, column_name):
            """Helper function to match columns case-insensitively."""
            return next((df_col for df_col in df.columns if df_col.lower() == column_name.lower()), None)

        # Find existing key columns in the DataFrame (case-insensitive)
        existing_key_columns = [match_column(df, col) for col in key_columns if match_column(df, col)]

        # Ensure all key columns have a value before proceeding
        for col in key_columns:
            if col not in new_data or new_data[col] == '':
                QMessageBox.warning(self, "Warning", f"Please enter a value for '{col}'.")
                return

        # Prepare the new tool data for comparison, only including key columns
        new_tool = {match_column(df, col): new_data[col] for col in key_columns if col in new_data and match_column(df, col)}

        # Stock Number validation
        if 'Stock Number' in new_data:
            stock_number = new_data['Stock Number']

            # Check if Stock Number exists and increment it from the last stock number
            if 'Stock Number' in df.columns:
                last_stock_number = df['Stock Number'].max()  # Get the last stock number
                if last_stock_number is not None:
                    expected_next_stock_number = self.increment_stock_number(last_stock_number)
                    if expected_next_stock_number != stock_number:
                        QMessageBox.warning(self, "Warning", f"The stock number must be {expected_next_stock_number}.")
                        return

            # Check for duplicate stock numbers
            if stock_number in df['Stock Number'].values:
                QMessageBox.warning(self, "Warning", "Tool with this stock number already exists.")
                return

        # Check for duplicate entries based on key columns
        if existing_key_columns:
            existing_tools = df[existing_key_columns].to_dict('records')

            for existing_tool in existing_tools:
                is_duplicate = True
                for col in existing_key_columns:
                    existing_value = existing_tool.get(col, '')  # Get the existing value, default to an empty string
                    new_value = new_tool.get(col, 0)  # Get the new value, default to 0

                    # Ensure both values are non-empty before comparing
                    if existing_value == '' or new_value == '':
                        is_duplicate = False
                        break

                    try:
                        # Convert both values to floats or integers for comparison, if possible
                        if isinstance(existing_value, (int, float)) or isinstance(new_value, (int, float)):
                            if float(existing_value) != float(new_value):
                                is_duplicate = False
                                break
                        else:
                            # Compare as strings if they are not numeric
                            if str(existing_value).strip() != str(new_value).strip():
                                is_duplicate = False
                                break
                    except ValueError:
                        is_duplicate = False
                        break

                if is_duplicate:
                    QMessageBox.warning(self, "Warning", "Tool already exists with the same specs.")
                    return

            # Append new tool to the dataframe if no duplicate found
            self.dataframes[tool] = df._append(new_data, ignore_index=True)
            
            # Add the stock number and material group to the List sheet
            list_df = self.dataframes.get("List", None)
            if list_df is not None:
                new_list_data = {
                    "Material": new_data["Stock Number"],
                    "Matl Group":matl_group
                }
                list_df = list_df._append(new_list_data, ignore_index=True)
                self.dataframes["List"] = list_df
            
            self.stock_number = new_data["Stock Number"]  # Set the stock number attribute
            QMessageBox.information(self, "Success", "Tool added successfully!")
            self.accept()
        else:
            QMessageBox.warning(self, "Warning", "Please enter values for the necessary columns.")



    def highlight_matching_rows(self, new_data, necessary_columns):
        self.table_view.clearSelection()  # Clear any existing selection
        
        for row in range(self.model.rowCount()):
            match = True
            for col in necessary_columns:
                try:
                    model_index = self.model.index(row, self.dataframes[self.tool_combo.currentText()].columns.get_loc(col))
                    cell_value = str(self.model.data(model_index)).lower()
                    if cell_value != new_data.get(col.lower(), ""):
                        match = False
                        break
                except:
                    match = False
                    break 

            if match:
                self.table_view.selectionModel().selectRow(row, QItemSelectionModel.SelectionFlag.Select)
                # Optionally, you can scroll to the first matched row:
                self.table_view.scrollTo(self.model.index(row, 0))


class ToolFinderApp(QMainWindow):
    """Main application class."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Modern Tool Finder")
        self.setGeometry(100, 100, 1200, 800)
        self.setup_ui()
        self.load_data()
        self.apply_dark_theme()
        self.authenticated = False
        self.auth_data = self.load_auth_data()
        

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        self.splitter = QSplitter(Qt.Orientation.Horizontal)

        self.filter_widget = QWidget()
        self.filter_layout = QGridLayout(self.filter_widget)
        self.splitter.addWidget(self.filter_widget)

        self.table_view = QTableView()
        self.table_view.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.setHorizontalScrollMode(
            QTableView.ScrollMode.ScrollPerPixel)
        self.table_view.setSortingEnabled(True)
        self.table_view.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.table_view.setSelectionBehavior(
            QTableView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(
            QTableView.SelectionMode.SingleSelection)
        self.table_view.setStyleSheet("background-color: white; color: black;")
        self.table_view.horizontalHeader().setStyleSheet("background-color: #2a2a2a; color: white;")
        self.table_view.verticalHeader().setStyleSheet("background-color: #2a2a2a; color: white;")
        self.splitter.addWidget(self.table_view)
        
        

        main_layout.addWidget(self.splitter)

        self.tool_combo = QComboBox()
        self.tool_combo.addItem("Select Tool")
        self.tool_combo.addItems([
            "Drills carbide", "Drills HSS", "Special Centre Drills", "Insert type Drills",
            "Drills", "FL Drill", "Countersink Drills", "GUN Drill", "Special Drill",
            "SPAD Drill", "Ball End Mills", "Flat End Mill", "Insert Type Cutter",
            "Toric End Mill", "Taper End Mill", "Slitting Wheel(Carbide)",
            "Chamfer Cutter(Carbide)", "Slitting Wheel(HSS)", "Special End Mill",
            "Batch End Mill", "T Slot Cutters", "Form Cutters", "Thread Form Cutters",
            "Thread Mill Relife Cutters", "HandTaps-MetricCoarse&Fineserie", "HandTaps UNC Series",
            "Hand Taps UNF Series", "Hand Taps UNJF", "Hand Taps Metric Fine Series",
            "Machine Taps Metric J Series", "Machine Taps UNJC,UNJF,UNJEF",
            "Machine Taps UNF & NF Series", "Machine Taps UNEF Series", "Machine Taps UNC Series",
            "Machine Taps UNS Type", "Helicoil Taps Metric&UN Series", "Helicoil Inserts Metric & UN",
            "Spiral Lock Taps", "HeliCoil Machine Taps", "HeliCoil Extracting tools",
            "Helicoil Hand Insert Tool (TANG)", "Helicoil Hand Install w|o TANG", "Rep.Inserty Blade Kit"
            "Rep.Insert Blade Kit", "Rep.Punch Tool", "TANG Removal Tool", "Tapping Adaptor",
            "Roll Taps", "Reamers", "Fixtures", "Inhouse"
        ])
        self.tool_combo.setCurrentIndex(0)
        self.tool_combo.currentTextChanged.connect(self.update_filter_widget)
        self.filter_layout.addWidget(QLabel("Tool:"), 0, 0)
        self.filter_layout.addWidget(self.tool_combo, 0, 1)
        self.tool_combo.setStyleSheet("background-color: #2a2a2a; color: white; selection-background-color: #3a92ea; selection-color: white;")

        self.matl_group_combo = QComboBox()
        self.matl_group_combo.addItem("Select Matl Group")
        self.matl_group_combo.addItems(
                ["ELECTRODE", "TOOL(PUR)", "TOOLS SPL", "TOOLFIX"])
        self.matl_group_combo.setCurrentIndex(0)
        self.matl_group_combo.currentTextChanged.connect(
                self.update_filter_widget)
        self.matl_group_combo.setEnabled(False)  # Initially disable matl_group_combo
        self.filter_layout.addWidget(QLabel("Material Group:"), 1, 0)
        self.filter_layout.addWidget(self.matl_group_combo, 1, 1)
        self.matl_group_combo.setStyleSheet("background-color: #2a2a2a; color: white; selection-background-color: #3a92ea; selection-color: white;")

        self.add_button = QPushButton("Add Tool")
        self.add_button.clicked.connect(self.add_tool)
        self.filter_layout.addWidget(self.add_button)

        self.delete_button = QPushButton("Delete Tool")
        self.delete_button.clicked.connect(self.delete_tool)
        self.filter_layout.addWidget(self.delete_button)

        self.update_button = QPushButton("Update Tool")
        self.update_button.clicked.connect(self.update_tool)
        self.filter_layout.addWidget(self.update_button)
        
        self.filter_widget_instance = None  # Connect the signal to the slot


    def apply_dark_theme(self):
        app = QApplication.instance()
        app.setStyle(QStyleFactory.create("Fusion"))

        dark_palette = QPalette()
        dark_palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
        dark_palette.setColor(
            QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
        dark_palette.setColor(QPalette.ColorRole.Base, QColor(35, 35, 35))
        dark_palette.setColor(
            QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
        dark_palette.setColor(
            QPalette.ColorRole.ToolTipBase, QColor(25, 25, 25))
        dark_palette.setColor(
            QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
        dark_palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
        dark_palette.setColor(QPalette.ColorRole.BrightText, Qt.GlobalColor.blue)
        dark_palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
        dark_palette.setColor(
            QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
        dark_palette.setColor(
            QPalette.ColorRole.BrightText, Qt.GlobalColor.red)
        dark_palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
        dark_palette.setColor(
            QPalette.ColorRole.Highlight, QColor(0, 255, 249))
        dark_palette.setColor(
            QPalette.ColorRole.HighlightedText, QColor(0,0,0))
        

        app.setPalette(dark_palette)

        # Set stylesheet for more vibrant buttons
        app.setStyleSheet("""
            QPushButton {
                background-color: #3a92ea;
                color: white;
                border: none;
                padding: 5px;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3a92ea;
            }
            QPushButton:pressed {
                background-color: #1a72ca;
            }
            QComboBox {
                background-color: #2a2a2a;
                color: white;
                border: 1px solid #3a3a3a;
                padding: 3px;
                border-radius: 3px;
            }
            /* Highlight the QComboBox when it's focused */
            QComboBox:focus {
                border: 2px solid #3a92ea;
                background-color: #3a3a3a;
                color: #ffffff;
            }
            
            /* Style for dropdown items */
            QComboBox QAbstractItemView {
                border: 1px solid #3a92ea;
                background-color: #2a2a2a;
                color: white;
                selection-background-color: #3a92ea;
                selection-color: white;
            }

            /* Highlight the dropdown label (title) when focused */
            QLabel:focus {
                color: #3a92ea;
            }

            /* Style for QLabel (dropdown title or label) */
            QLabel {
                color: white;
                font-weight: bold;
                padding: 2px;
            }
             
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 15px;
                border-left-width: 1px;
                border-left-color: #3a3a3a;
                border-left-style: solid;
                image: url('arrow_down.png');  # Replace with your arrow image path
            }
            /* Table view settings */
            QTableView {
                background-color: #2a2a2a;
                alternate-background-color: #3a3a3a;
                color: white;
            }
            QHeaderView::section {
                background-color: #2a2a2a;
                color: white;
                padding: 5px;
                border: 1px solid #3a3a3a;
            }
            QTableView::item:selected {
                background-color: #3a92ea;
            }
        """)

    def on_header_clicked(self, logical_index):
        if self.table_view.model():
            order = Qt.SortOrder.AscendingOrder if self.table_view.horizontalHeader().sortIndicatorSection() != logical_index or \
                self.table_view.horizontalHeader().sortIndicatorOrder() == Qt.SortOrder.DescendingOrder else Qt.SortOrder.DescendingOrder
            self.table_view.model().sort(logical_index, order)

    def load_data(self):
        try:
            self.file_path = "c:/Users/sugan/Downloads/tooling.xlsx"
            self.sheets = pd.read_excel(self.file_path, sheet_name=None)
            self.dataframes = {}
            for sheet_name, data in self.sheets.items():
                data.columns = data.columns.str.strip()
                self.dataframes[sheet_name] = pd.DataFrame(data)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error loading data: {e}")

    def update_filter_widget(self):
        tool = self.tool_combo.currentText()
        matl_group = self.matl_group_combo.currentText()

        if tool != "Select Tool":
            self.matl_group_combo.setEnabled(True)  # Enable matl_group_combo when a tool is selected
        else:
            self.matl_group_combo.setEnabled(False)  # Disable matl_group_combo when no tool is selected

        if tool != "Select Tool" or matl_group != "Select Matl Group":
            df = self.dataframes.get(tool, None)
            if df is not None:
                if hasattr(self, 'filter_widget_instance') and self.filter_widget_instance is not None:
                    self.filter_layout.removeWidget(self.filter_widget_instance)
                    self.filter_widget_instance.deleteLater()
                    self.filter_widget_instance = None

                if matl_group != "Select Matl Group":
                    list_df = self.dataframes.get("List", None)
                    if list_df is not None:
                        filtered_list_df = list_df[list_df['Matl Group'] == matl_group]
                        df = df[df['Stock Number'].isin(filtered_list_df['Material'])]

                self.filter_widget_instance = FilterWidget(parent=self, columns=df.columns.tolist(), data=df)
                self.filter_layout.addWidget(self.filter_widget_instance, 5, 0, 1, 2)
                self.filter_widget_instance.filtersApplied.connect(self.apply_filters)
                self.filter_widget_instance.resetFilters.connect(self.reset_all_filters)
                self.add_button.hide()
            else:
                QMessageBox.warning(self, "Warning", f"No data available for {tool}.")
                self.add_button.show()
        else:
            if hasattr(self, 'filter_widget_instance') and self.filter_widget_instance is not None:
                self.filter_layout.removeWidget(self.filter_widget_instance)
                self.filter_widget_instance.deleteLater()
                self.filter_widget_instance = None
            self.apply_filters({})
            self.add_button.show()

    def reset_all_filters(self):
        self.tool_combo.setCurrentIndex(0)
        self.matl_group_combo.setCurrentIndex(0)
        self.matl_group_combo.setEnabled(False)
        if hasattr(self, 'filter_widget_instance') and self.filter_widget_instance is not None:
            self.filter_layout.removeWidget(self.filter_widget_instance)
            self.filter_widget_instance.deleteLater()
            self.filter_widget_instance = None
        self.apply_filters({})

    def apply_filters(self, criteria):
        tool = self.tool_combo.currentText()
        matl_group = self.matl_group_combo.currentText()

        df = self.dataframes.get(tool, None)
        list_df = self.dataframes.get("List", None)

        if df is not None and list_df is not None:
            # Drop the 'Matl Group' column from the 'df' dataframe
            if 'Matl Group' in df.columns:
                df = df.drop('Matl Group', axis=1)

            # Merge the tool dataframe with the List dataframe
            merged_df = pd.merge(df, list_df[['Material', 'Matl Group']], left_on='Stock Number', right_on='Material', how='left')

            # If a specific Material Group is selected, filter by it
            if matl_group != "Select Matl Group":
                merged_df = merged_df[merged_df['Matl Group'] == matl_group]

            # Apply additional filters
            for col, value in criteria.items():
                if isinstance(value, tuple):
                    selected_value, selected_range = value
                    if selected_value and selected_value != "Select Value":
                        try:
                            selected_value = float(selected_value)
                            if selected_range and selected_range != "Select Range":
                                range_value = float(selected_range.replace('±', ''))
                                merged_df = merged_df[(merged_df[col] >= selected_value - range_value) & (merged_df[col] <= selected_value + range_value)]
                            else:
                                merged_df = merged_df[merged_df[col] == selected_value]
                        except ValueError:
                            continue
                elif value and value != "Select Value":
                    if pd.api.types.is_string_dtype(merged_df[col]):
                        merged_df = merged_df[merged_df[col].astype(str).str.lower() == str(value).lower()]
                    else:
                        try:
                            merged_df = merged_df[merged_df[col] == int(value)]  # Try to convert to int
                        except ValueError:
                            try:
                                merged_df = merged_df[merged_df[col] == float(value)]  # If not int, try to convert to float
                            except ValueError:
                                pass

            # Check if 'Matl Group' is present in the merged_df columns
            if 'Matl Group' in merged_df.columns:
                # Rearrange columns to have 'Matl Group' right after 'Stock Number'
                columns = merged_df.columns.tolist()
                matl_group_index = columns.index('Matl Group')
                stock_number_index = columns.index('Stock Number')
                columns.remove('Matl Group')
                columns.insert(stock_number_index + 1, 'Matl Group')

                # Remove the 'Material' column as it's redundant with 'Stock Number'
                if 'Material' in columns:
                    columns.remove('Material')

                merged_df = merged_df[columns]
            else:
                # If 'Matl Group' is not present, just remove the 'Material' column
                if 'Material' in merged_df.columns:
                    merged_df = merged_df.drop('Material', axis=1)

            print(f"Filtered Data Shape: {merged_df.shape}")
            model = PandasModel(merged_df)
            self.table_view.setModel(model)
            self.table_view.sortByColumn(0, Qt.SortOrder.AscendingOrder)
        else:
            if tool != "Select Tool":
                QMessageBox.warning(self, "Warning", f"No data available for {tool}.")
            self.table_view.setModel(None)


    def save_data(self):
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            for sheet_name, data in self.dataframes.items():
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet.delete_rows(1, sheet.max_row)
                else:
                    sheet = workbook.create_sheet(sheet_name)

                for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

            workbook.save(self.file_path)
            workbook.close()

            self.load_data()
            QMessageBox.information(
                self, "Success", "Data saved successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving data: {str(e)}")
    
    def load_auth_data(self):
        try:
            workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet = workbook["Auth"]
            sheet.sheet_state = "visible"  # Make the sheet visible temporarily
            auth_data = {}
            for row in range(2, sheet.max_row + 1):
                username = sheet.cell(row=row, column=1).value
                hashed_password = sheet.cell(row=row, column=2).value
                auth_data[username] = hashed_password
            sheet.sheet_state = "hidden"  # Hide the sheet again
            return auth_data
        except Exception as e:
            print(f"Error loading auth data: {e}")
            return None

    def authenticate(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Authentication")
        layout = QFormLayout()
        dialog.setLayout(layout)

        username_label = QLabel("Username:")
        self.username_input = QLineEdit()
        layout.addRow(username_label, self.username_input)

        password_label = QLabel("Password:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow(password_label, self.password_input)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            username = self.username_input.text().strip()  # Strip spaces
            password = self.password_input.text().strip()  # Strip spaces

            print(f"Entered username: '{username}'")

            # Load the workbook for logging
            try:
                log_workbook = openpyxl.load_workbook("C:/Users/sugan/Downloads/logs.xlsx")
            except FileNotFoundError:
                log_workbook = Workbook()
                log_workbook.save("C:/Users/sugan/Downloads/logs.xlsx")
                log_workbook = openpyxl.load_workbook("C:/Users/sugan/Downloads/logs.xlsx")

            # Create a new sheet for logging if it doesn't exist
            if "Log" not in log_workbook.sheetnames:
                log_workbook.create_sheet("Log")
                log_sheet = log_workbook["Log"]
                log_sheet['A1'] = "Timestamp"
                log_sheet['B1'] = "Username"
                log_sheet['C1'] = "Operation"
                log_sheet['D1'] = "Result"
            else:
                log_sheet = log_workbook["Log"]

            # Log the authentication attempt
            log_row = [datetime.datetime.now(), username, "Login", "Attempted"]
            log_sheet.append(log_row)

            # Check the username and password against the auth data
            if username in self.auth_data:
                hashed_password = self.hash_password(password)  # Hash the entered password
                stored_hashed_password = self.auth_data[username].strip()  # Get the stored hashed password

                # Compare hashed passwords
                if self.compare_passwords(hashed_password, stored_hashed_password):
                    print("Authentication successful!")
                    self.authenticated = True
                    self.authenticated_username = username  # Store the username in a class attribute
                    # Log the successful authentication
                    log_row = [datetime.datetime.now(), username, "Login", "Successful"]
                    log_sheet.append(log_row)

                    # Save the log to the Excel file
                    log_workbook.save("C:/Users/sugan/Downloads/logs.xlsx")

                    return True  # Return True if authentication is successful
                else:
                    print("Password mismatch!")
                    QMessageBox.warning(self, "Authentication Failed", "Invalid username or password")

                    # Log the failed authentication
                    log_row = [datetime.datetime.now(), username, "Login", "Failed"]
                    log_sheet.append(log_row)

                    # Save the log to the Excel file
                    log_workbook.save("C:/Users/sugan/Downloads/logs.xlsx")
            else:
                print("Username not found!")
                QMessageBox.warning(self, "Authentication Failed", "Invalid username or password")

                # Log the failed authentication
                log_row = [datetime.datetime.now(), username, "Login", "Failed"]
                log_sheet.append(log_row)

                # Save the log to the Excel file
                log_workbook.save("C:/Users/sugan/Downloads/logs.xlsx")
        else:
            self.authenticated = False

        return False  # Return False if authentication fails

    def log_operation(self, operation, stock_number=None):
        try:
            log_workbook = openpyxl.load_workbook("C:/Users/sugan/Downloads/logs.xlsx")
        except FileNotFoundError:
            log_workbook = Workbook()
            log_workbook.save("C:/Users/sugan/Downloads/authentication_log.xlsx")
            log_workbook = openpyxl.load_workbook("C:/Users/sugan/Downloads/logs.xlsx")

        log_sheet = log_workbook["Log"]
        username = self.authenticated_username  # Use the stored username
        log_row = [datetime.datetime.now(), username, operation, "Attempted"]
        if stock_number:
            log_row[2] = f"{operation} ({stock_number})"
        log_sheet.append(log_row)
        log_workbook.save("C:/Users/sugan/Downloads/logs.xlsx")

    def add_tool(self):
        if not self.authenticate():  # Authenticate and check if the user is authenticated
            return  # If authentication fails, exit the method

        dialog = AddToolDialog(self.dataframes, self)
        if dialog.exec():  # Show the dialog and wait for it to close
            stock_number = dialog.stock_number  # Get the stock number attribute
            self.log_operation("Add Tool", stock_number)  # Log the operation with the stock number
            self.save_data()  # Save data only if the dialog is accepted
            self.update_filter_widget()  # Update the filter widget after saving

    def update_tool(self):
        # Prompt for authentication each time the update_tool method is called
        if not self.authenticate():  # Call authenticate and check if the user is authenticated
            return

        tool = self.tool_combo.currentText()
        df = self.dataframes.get(tool, None)
        list_df = self.dataframes.get("List", None)

        if df is not None and list_df is not None:
            selected_rows = self.table_view.selectionModel().selectedRows()
            if selected_rows:
                view_index = selected_rows[0].row()
                model_index = self.table_view.model().index(view_index, 0)
                stock_number = self.table_view.model().data(model_index)

                original_index = df[df['Stock Number'] == stock_number].index[0]
                old_data = df.iloc[original_index].to_dict()

                dialog = QDialog(self)
                dialog.setWindowTitle("Update Tool")
                layout = QFormLayout()
                dialog.setLayout(layout)

                inputs = {}
                for col, value in old_data.items():
                    if col in ['drill dia', 'flute length', 'corner radius', 'Cutter dia']:
                        input_widget = QDoubleSpinBox()
                        input_widget.setValue(float(value))
                    else:
                        input_widget = QLineEdit()
                        input_widget.setText(str(value))
                    layout.addRow(QLabel(col), input_widget)
                    inputs[col] = input_widget

                button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addRow(button_box)

                if dialog.exec() == QDialog.DialogCode.Accepted:
                    for col, input_widget in inputs.items():
                        if col in ['drill dia', 'flute length', 'corner radius', 'Cutter dia']:
                            df.at[original_index, col] = input_widget.value()
                        else:
                            df.at[original_index, col] = input_widget.text()

                    # Update the List sheet only if the stock number is updated
                    if inputs['Stock Number'].text() != stock_number:
                        list_index = list_df[list_df['Material'] == stock_number].index[0]
                        list_df.at[list_index, 'Material'] = inputs['Stock Number'].text()

                    self.dataframes[tool] = df
                    self.dataframes["List"] = list_df
                    self.save_data()
                    self.apply_filters(self.filter_widget_instance.get_filter_criteria())
                    self.log_operation("Update Tool", stock_number)
                    QMessageBox.information(self, "Success", f"Tool {stock_number} updated successfully!")
            else:
                QMessageBox.warning(self, "Warning", "No row selected for updating.")
        else:
            QMessageBox.warning(self, "Warning", "No tool selected.")

                
    def delete_tool(self):
        if not self.authenticate():  # Call authenticate and check if the user is authenticated
            return

        tool = self.tool_combo.currentText()
        df = self.dataframes.get(tool, None)
        list_df = self.dataframes.get("List", None)

        if df is not None and list_df is not None:
            selected_rows = self.table_view.selectionModel().selectedRows()
            if selected_rows:
                stock_numbers = []
                for row in selected_rows:
                    view_index = row.row()
                    model_index = self.table_view.model().index(view_index, 0)
                    stock_number = self.table_view.model().data(model_index)
                    stock_numbers.append(stock_number)

                reply = QMessageBox.question(self, 'Confirm Deletion',
                                            f"Are you sure you want to delete {len(stock_numbers)} selected tool(s)?\n"
                                            "This will remove them from both the specific tool sheet and the List sheet.",
                                            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                            QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes:
                    self.dataframes[tool] = df[~df['Stock Number'].isin(stock_numbers)].reset_index(drop=True)
                    self.dataframes["List"] = list_df[~list_df['Material'].isin(stock_numbers)].reset_index(drop=True)
                    self.save_data()
                    self.apply_filters(self.filter_widget_instance.get_filter_criteria())
                    self.log_operation("Delete Tool", stock_numbers[0])
                    QMessageBox.information(self, "Success",
                                            f"{len(stock_numbers)} tool(s) deleted successfully from both sheets!")
            else:
                QMessageBox.warning(self, "Warning", "No rows selected for deletion.")
        else:
            QMessageBox.warning(self, "Warning", "No tool selected or List sheet not found.")


    def hash_password(self, password):
        # Use a secure hashing algorithm like SHA-256
        return hashlib.sha256(password.encode('utf-8')).hexdigest()

    def compare_passwords(self, hashed_password1, hashed_password2):
        # Use a secure comparison function like hmac.compare_digest
        return hmac.compare_digest(hashed_password1, hashed_password2)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = ToolFinderApp()
    main_window.show()
    sys.exit(app.exec())
